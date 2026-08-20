import ast
import csv
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NPU_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "npu"))
DOCS_DIR = os.path.dirname(os.path.abspath(__file__))

DTYPE_RE = re.compile(
    r"\b(fp32|fp16|bf16|fp64|f32|f16|f64|int8|int16|int32|int64|uint32|uint64|bool|float32|float64|float16|bfloat16)\b"
)


def _dedupe(items):
    seen = set()
    out = []
    for it in items:
        if it not in seen:
            seen.add(it)
            out.append(it)
    return out


def _first_arg(argstr):
    depth = 0
    first = []
    for ch in argstr:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        elif ch == "," and depth == 0:
            break
        first.append(ch)
    return "".join(first).strip()


def _split_name_shape(name):
    name = name.strip().rstrip(",")
    m = re.match(r"([A-Za-z_]\w*)\s*(\[[^\]]*\])", name)
    if m:
        return m.group(1), m.group(2)
    return name, ""


def get_header(src):
    lines = []
    for line in src.splitlines():
        s = line.strip()
        if s.startswith("#"):
            lines.append(s.lstrip("# ").strip())
        elif not s:
            continue
        else:
            break
    return lines


def extract_kernel_signature(src):
    """从文档字符串的 Kernel signature 块提取带 [shape] 注释的参数（精简、去重）。"""
    m = re.search(r'"""(.*?)"""', src, re.DOTALL)
    if not m:
        return []
    started = False
    raw = []
    for line in m.group(1).splitlines():
        ls = line.strip()
        if "(" in ls and "_kernel" in ls:
            started = True
            continue
        if not started:
            continue
        if ls.endswith(")"):
            break
        raw.append(line)

    sig = []
    for line in raw:
        if "#" not in line:
            name, shape = _split_name_shape(line)
            if shape:
                sig.append(f"{name} {shape}")
            continue
        name = line.split(",")[0].strip()
        comment = line.split("#", 1)[1].strip()
        if "stride" in name or "constexpr" in name:
            continue
        mshape = re.search(r"\[[^\]]*\]", comment)
        if not mshape:
            continue
        shape = mshape.group(0)
        dm = DTYPE_RE.search(comment)
        dtype = dm.group(1) if dm else ""
        entry = f"{name} {shape}"
        if dtype:
            entry += f" {dtype}"
        sig.append(entry)
    return _dedupe(sig)


def extract_tensor_assigns(src):
    """从 torch 张量创建语句提取 shape 与 dtype（精简、去重）。"""
    pat = re.compile(r"\b(\w+)\s*=\s*torch\.(\w+)\(((?:[^()]|\([^()]*\))*)\)")
    out = []
    for name, fn, argstr in pat.findall(src):
        argstr = argstr.strip()
        if fn not in ("zeros", "ones", "full", "empty", "randn", "rand", "randint"):
            continue
        dm = re.search(r"dtype\s*=\s*torch\.(\w+)", argstr)
        dtype = dm.group(1) if dm else ""

        shape = ""
        km = re.search(r"(?:size|shape)\s*=\s*(\([^)]*\)|\w+)", argstr)
        if km:
            shape = km.group(1)
        elif fn in ("zeros", "ones", "full", "empty", "randn", "rand"):
            shape = _first_arg(argstr)
        # randint 无 size= 时首参是 low，无法确定 shape，跳过
        if not shape:
            continue
        shape = re.sub(r"\s+", "", shape).strip("()")
        if (
            not shape
            or re.search(r"dtype|device|low|high|=|-inf|DEVICE|_shape", shape)
        ):
            continue
        seg = f"{name}[{shape}]"
        if dtype:
            seg += f" {dtype}"
        out.append(seg)
    return _dedupe(out)


def extract_module_constants(tree):
    consts = {}
    for node in tree.body:
        if isinstance(node, ast.Assign):
            for t in node.targets:
                if isinstance(t, ast.Name) and t.id.isupper():
                    try:
                        v = ast.literal_eval(node.value)
                        if isinstance(v, (int, float, str)):
                            consts[t.id] = v
                    except Exception:
                        pass
    return consts


def extract_imports(tree):
    imp = []
    for node in ast.walk(tree):
        if isinstance(node, ast.ImportFrom) and node.module and "vllm" in node.module:
            names = [a.name for a in node.names]
            keep = [n for n in names if n not in ("tl", "triton") and not n.startswith("init_")]
            imp.append((node.module, keep))
    return imp


def extract_parametrize(tree):
    params = []
    for node in ast.walk(tree):
        if isinstance(node, ast.FunctionDef):
            for dec in node.decorator_list:
                call = dec if isinstance(dec, ast.Call) else (
                    dec.value if isinstance(dec, ast.Attribute) and isinstance(dec.value, ast.Call) else None
                )
                if call is None:
                    continue
                func = call.func
                fname = func.id if isinstance(func, ast.Name) else (
                    func.attr if isinstance(func, ast.Attribute) else None
                )
                if fname != "parametrize":
                    continue
                if len(call.args) >= 2 and isinstance(call.args[0], ast.Constant):
                    argname = call.args[0].value
                    try:
                        vals = ast.unparse(call.args[1])
                    except Exception:
                        vals = "?"
                    params.append((argname, vals))
    return params


def extract_test_funcs(tree):
    return [n.name for n in ast.walk(tree) if isinstance(n, ast.FunctionDef) and n.name.startswith("test_")]


def _norm_location(src):
    m = re.search(r"# Kernel source:\s*(.+)", src)
    if m:
        return m.group(1).strip()
    return ""


def main():
    rows = []
    for fn in sorted(os.listdir(NPU_DIR)):
        if not fn.endswith(".py") or fn.startswith("__"):
            continue
        path = os.path.join(NPU_DIR, fn)
        with open(path, encoding="utf-8") as f:
            src = f.read()
        try:
            tree = ast.parse(src)
        except SyntaxError:
            continue

        header = get_header(src)
        coverage = ""
        kernel_source = ""
        for h in header:
            if h.startswith("Coverage:"):
                coverage = h[len("Coverage:"):].strip()
            elif h.startswith("Kernel source:"):
                kernel_source = h[len("Kernel source:"):].strip()

        sig = extract_kernel_signature(src)
        assigns = extract_tensor_assigns(src)
        consts = extract_module_constants(tree)
        params = extract_parametrize(tree)
        tests = extract_test_funcs(tree)

        op = coverage or consts.get("OPERATOR", "") or os.path.splitext(fn)[0].replace("test_", "")
        op = op.split(" via ")[0].strip()

        loc = kernel_source
        if not loc:
            for m, names in extract_imports(tree):
                if "triton_utils" in m or "runtime_npu" in m:
                    continue
                if names:
                    loc = m
                    break

        main_input = "; ".join(sig) if sig else "; ".join(assigns) if assigns else ""
        is_unwired = "npu_upstream_unwired" in src or "npu_adaptation_status" in src
        if is_unwired:
            main_input = "无 NPU 适配（unwired，skip）"

        param_str = " | ".join(dict.fromkeys(f"{k}={v}" for k, v in params))
        param_str = " | ".join(_dedupe(param_str.split(" | ")))
        const_str = "; ".join(f"{k}={v}" for k, v in sorted(consts.items()))

        rows.append([op, loc, main_input, ", ".join(tests), param_str, const_str])

    cols = ["算子", "位置", "主输入 shape", "测试函数", "参数化（参数=取值）", "固定 shape"]

    # 精简：若某算子主输入过于冗长，截断保留前若干个唯一项
    cleaned = []
    for r in rows:
        parts = [p.strip() for p in r[2].split(";") if p.strip()]
        r[2] = "; ".join(parts[:12]) + (" …" if len(parts) > 12 else "")
        cleaned.append(r)
    rows = cleaned

    out_xlsx = os.path.join(DOCS_DIR, "npu_算子shape汇总.xlsx")
    write_xlsx(cols, rows, out_xlsx)

    out_csv = os.path.join(DOCS_DIR, "npu_算子shape汇总.csv")
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(cols)
        w.writerows(rows)

    out_md = os.path.join(DOCS_DIR, "npu_算子shape汇总.md")
    with open(out_md, "w", encoding="utf-8") as f:
        f.write("# strict_ut NPU 算子 shape 汇总\n\n")
        f.write("| " + " | ".join(cols) + " |\n")
        f.write("|" + " --- |" * len(cols) + "\n")
        for r in rows:
            cells = [c.replace("\n", " ").replace("|", "/") for c in r]
            f.write("| " + " | ".join(cells) + " |\n")

    print("WROTE", out_xlsx, "rows=", len(rows))


def write_xlsx(cols, rows, out_xlsx):
    wb = Workbook()
    ws = wb.active
    ws.title = "算子shape汇总"

    ws.append(cols)
    header_fill = PatternFill("solid", fgColor="4472C4")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append(r)

    widths = [28, 42, 70, 46, 46, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_xlsx)


if __name__ == "__main__":
    main()